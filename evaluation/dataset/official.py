"""Verification against the OFFICIAL local KArSL-502 label workbook.

The workbook shipped with the downloaded dataset is the authoritative label
source. Public mirrors, candidate CSVs and inferred SignID ranges are never
preferred over it. Values are preserved exactly as stored; Unicode NFC
normalization is used for equality comparison only and never written back.
"""

from __future__ import annotations

import hashlib
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from .core28 import CORE28_SIGN_IDS, EXTENDED_LETTER_SIGN_IDS, LabelRecord

# Frozen once the official workbook has confirmed the mapping.
OFFICIAL_MAPPING_VERSION = "karsl-core28-v2-official"

# The workbook's own chapter structure, discovered rather than assumed:
# SignIDs 1-31 are numeric magnitudes, 32-59 the 28 standard alphabet letters,
# 60-70 extended letter forms. The KArSL filename convention encodes the same
# split in its leading field ("01" for 1-31, "02" for 32-70), which is an
# independent corroboration.
NUMBER_SIGN_IDS: tuple[int, ...] = tuple(range(1, 32))
LETTER_SIGN_IDS: tuple[int, ...] = tuple(range(32, 71))

CATEGORY_NUMBER = "number"
CATEGORY_CORE28_LETTER = "core28_letter"
CATEGORY_EXTENDED_LETTER = "extended_letter"
CATEGORY_OTHER = "other"


def normalize_for_comparison(value: object) -> str:
    """NFC-normalized, stripped text used ONLY for equality checks."""

    return unicodedata.normalize("NFC", str(value)).strip()


def sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for block in iter(lambda: handle.read(1 << 20), b""):
            digest.update(block)
    return digest.hexdigest()


@dataclass(frozen=True, slots=True)
class WorkbookEntry:
    """One row of the official workbook, preserved verbatim."""

    sign_id: int
    label_ar: str
    label_en: str
    source_row: int
    arabic_cell_type: str
    category: str


@dataclass(slots=True)
class WorkbookVerification:
    path: str
    file_name: str
    size_bytes: int
    sha256: str
    sheet_names: list[str]
    sheet_used: str
    total_rows: int
    header: list[str]
    data_rows: int
    sign_id_min: int
    sign_id_max: int
    entries: list[WorkbookEntry] = field(default_factory=list)
    issues: list[str] = field(default_factory=list)

    def by_id(self) -> dict[int, WorkbookEntry]:
        return {entry.sign_id: entry for entry in self.entries}


def _categorize(sign_id: int) -> str:
    if sign_id in CORE28_SIGN_IDS:
        return CATEGORY_CORE28_LETTER
    if sign_id in EXTENDED_LETTER_SIGN_IDS:
        return CATEGORY_EXTENDED_LETTER
    if sign_id in NUMBER_SIGN_IDS:
        return CATEGORY_NUMBER
    return CATEGORY_OTHER


def read_official_workbook(path: str | Path) -> WorkbookVerification:
    """Read and characterize the official workbook without altering it.

    Numeric SignIDs (the number/digit chapter) are stored as spreadsheet
    integers rather than text. They are rendered with ``str`` and the original
    cell type is recorded, so nothing is silently reinterpreted as an
    Arabic-Indic numeral form the workbook does not actually contain.
    """

    import openpyxl

    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Official label workbook not found: {workbook_path}")

    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_names = list(workbook.sheetnames)
    sheet = workbook[sheet_names[0]]
    rows = [row for row in sheet.iter_rows(values_only=True)]
    if not rows:
        raise ValueError(f"Official workbook {workbook_path} is empty")

    header = [str(cell) if cell is not None else "" for cell in rows[0]]
    issues: list[str] = []
    if header[:3] != ["SignID", "Sign-Arabic", "Sign-English"]:
        issues.append(f"unexpected header: {header}")

    entries: list[WorkbookEntry] = []
    seen: set[int] = set()
    for offset, row in enumerate(rows[1:], start=2):
        if row[0] is None:
            continue
        try:
            sign_id = int(row[0])
        except (TypeError, ValueError):
            issues.append(f"row {offset}: non-integer SignID {row[0]!r}")
            continue
        if sign_id in seen:
            issues.append(f"row {offset}: duplicate SignID {sign_id}")
            continue
        seen.add(sign_id)
        arabic = row[1]
        english = row[2] if len(row) > 2 else None
        entries.append(
            WorkbookEntry(
                sign_id=sign_id,
                label_ar="" if arabic is None else str(arabic),
                label_en="" if english is None else str(english),
                source_row=offset,
                arabic_cell_type=type(arabic).__name__,
                category=_categorize(sign_id),
            )
        )

    sign_ids = sorted(seen)
    expected = list(range(sign_ids[0], sign_ids[-1] + 1)) if sign_ids else []
    if sign_ids != expected:
        issues.append(f"SignIDs are not contiguous over {sign_ids[0]}..{sign_ids[-1]}")

    return WorkbookVerification(
        path=str(workbook_path),
        file_name=workbook_path.name,
        size_bytes=workbook_path.stat().st_size,
        sha256=sha256_file(workbook_path),
        sheet_names=sheet_names,
        sheet_used=sheet_names[0],
        total_rows=len(rows),
        header=header,
        data_rows=len(entries),
        sign_id_min=sign_ids[0] if sign_ids else 0,
        sign_id_max=sign_ids[-1] if sign_ids else 0,
        entries=entries,
        issues=issues,
    )


def verify_candidate_mapping(
    verification: WorkbookVerification, candidate: list[LabelRecord], expected_ids: tuple[int, ...]
) -> dict[str, Any]:
    """Check a candidate mapping against the official workbook.

    Returns a PASS/FAIL result plus every disagreement. A disagreement is a
    reason to correct the candidate, never a reason to rewrite the workbook.
    """

    official = verification.by_id()
    candidate_ids = tuple(record.sign_id for record in candidate)
    mismatches: list[dict[str, Any]] = []
    missing = [sign_id for sign_id in expected_ids if sign_id not in official]

    for record in candidate:
        entry = official.get(record.sign_id)
        if entry is None:
            continue
        if normalize_for_comparison(record.label_ar) != normalize_for_comparison(entry.label_ar):
            mismatches.append({
                "sign_id": record.sign_id, "field": "label_ar",
                "candidate": record.label_ar, "official": entry.label_ar,
            })
        candidate_en = record.label_en or ""
        if normalize_for_comparison(candidate_en) != normalize_for_comparison(entry.label_en):
            mismatches.append({
                "sign_id": record.sign_id, "field": "label_en",
                "candidate": candidate_en, "official": entry.label_en,
            })

    ids_match = candidate_ids == expected_ids
    return {
        "expected_sign_ids": list(expected_ids),
        "candidate_sign_ids": list(candidate_ids),
        "sign_ids_match": ids_match,
        "class_count": len(expected_ids),
        "missing_from_workbook": missing,
        "label_mismatches": mismatches,
        "passed": bool(ids_match and not missing and not mismatches),
    }


def category_breakdown(verification: WorkbookVerification) -> dict[str, Any]:
    """Number / Core-28 / extended-letter classification straight from the workbook."""

    buckets: dict[str, list[WorkbookEntry]] = {}
    for entry in verification.entries:
        if entry.sign_id > max(LETTER_SIGN_IDS):
            continue
        buckets.setdefault(entry.category, []).append(entry)

    def describe(name: str) -> dict[str, Any]:
        rows = sorted(buckets.get(name, []), key=lambda item: item.sign_id)
        return {
            "class_count": len(rows),
            "sign_ids": [row.sign_id for row in rows],
            "entries": [
                {
                    "sign_id": row.sign_id,
                    "label_ar": row.label_ar,
                    "label_en": row.label_en,
                    "arabic_cell_type": row.arabic_cell_type,
                }
                for row in rows
            ],
        }

    return {
        CATEGORY_NUMBER: describe(CATEGORY_NUMBER),
        CATEGORY_CORE28_LETTER: describe(CATEGORY_CORE28_LETTER),
        CATEGORY_EXTENDED_LETTER: describe(CATEGORY_EXTENDED_LETTER),
    }
